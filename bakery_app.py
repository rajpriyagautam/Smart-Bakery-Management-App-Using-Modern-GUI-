# sweet_bakery_pro_v5.py
import customtkinter as ctk
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from tkinter import messagebox
import datetime
from pathlib import Path
BASE_DIR = Path(__file__).parent
ASSETS_FOLDER = BASE_DIR / "assets"
# ---------------------- Config ----------------------
ctk.set_appearance_mode("light")

ITEMS = [
    ("Cake", ASSETS_FOLDER / "cake.png", 250),
    ("Pastries", ASSETS_FOLDER / "pastries.png", 35),
    ("Bread", ASSETS_FOLDER / "bread.png", 40),
    ("Cookies", ASSETS_FOLDER / "cookies.png", 15),
    ("Chocolate", ASSETS_FOLDER / "chocolate.png", 70),
    ("Donut", ASSETS_FOLDER / "donut.png", 60),
    ("Cupcake", ASSETS_FOLDER / "cupcake.png", 25),
    ("Candies", ASSETS_FOLDER / "candies.png", 5),
]

BILLS_FOLDER = Path("Bakery Bills")
BILLS_FOLDER.mkdir(parents=True, exist_ok=True)
OWNER_FILE = BILLS_FOLDER / "Owner_Records.xlsx"
LOGO_FILE = ASSETS_FOLDER / "logo.png"

# ---------------------- App ----------------------
root = ctk.CTk()
root.geometry("1100x780")
root.title("🍰 Sweet Bakery ")
root.configure(fg_color="#F2ECE0")

# ---------------------- Utils ----------------------
def safe_filename(name):
    return "".join(c for c in name if c.isalnum() or c in (" ", "_")).strip().replace(" ", "_") or "Guest"

def ensure_excel_headers(path):
    if not path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Records"
        ws.append(["Customer Name", "Date & Time", "Total Bill"])
        wb.save(path)


def append_owner_record(customer, total_bill):
    ensure_excel_headers(OWNER_FILE)

    wb = load_workbook(OWNER_FILE)
    ws = wb.active

    # Just 1 entry per order:
    ws.append([
        customer,
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        total_bill
    ])

    wb.save(OWNER_FILE)

    

def append_customer_file(customer, items):
    file_path = BILLS_FOLDER / f"{safe_filename(customer)}.xlsx"

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Bill"

    # Styles
    bold_center = Font(bold=True, size=12)
    center = Alignment(horizontal="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 🎂 Header Section
    ws.merge_cells("A1:D1")
    ws["A1"] = "🎂 SWEET BAKERY 🎂"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = "Customer Bill Receipt"
    ws["A2"].font = Font(bold=True, size=14)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Customer Name:"
    ws["B3"] = customer
    ws["A4"] = "Date & Time:"
    ws["B4"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Column headers
    ws.append(["Item", "Quantity", "Price", "Cost"])
    header_row = ws.max_row
    for col in range(1, 5):
        cell = ws.cell(row=header_row, column=col)
        cell.font = bold_center
        cell.alignment = center
        cell.border = border

    # Data rows
    total = 0
    for n, q, p, cost, grand in items:
        ws.append([n, q, p, cost])
        total += cost
        row = ws.max_row
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.alignment = center
            cell.border = border

    # Total row
    ws.append(["", "", "TOTAL", total])
    total_row = ws.max_row
    for col in range(1, 5):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
        cell.alignment = center
        cell.border = border

    # Thank-you note
    ws.merge_cells(f"A{total_row+2}:D{total_row+2}")
    ws[f"A{total_row+2}"] = "🙏 THANK YOU FOR COMING 🙏"
    ws[f"A{total_row+2}"].font = Font(bold=True, size=12)
    ws[f"A{total_row+2}"].alignment = Alignment(horizontal="center")

    # Adjust column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14

    wb.save(file_path)
    return file_path

# ---------------------- State ----------------------
order_state = {
    "customer": "Guest",
    "quantities": {n: 0 for n, _, _ in ITEMS},
    "selected": {n: ctk.BooleanVar(value=False) for n, _, _ in ITEMS},
}
qty_labels = {}

# ---------------------- Nav ----------------------
def show(frame):
    for w in root.winfo_children():
        w.pack_forget()
    frame.pack(expand=True, fill="both", padx=15, pady=15)

# ---------------------- Logo ----------------------
logo_ctk = None
if LOGO_FILE.exists():
    try:
        pil = Image.open(LOGO_FILE).convert("RGBA").resize((100, 100))
        logo_ctk = ctk.CTkImage(light_image=pil, dark_image=pil, size=(100, 100))
    except Exception:
        pass

# ---------------------- Page 1 ----------------------
welcome = ctk.CTkFrame(root, fg_color="#FFF8E1", corner_radius=16)
if logo_ctk:
    ctk.CTkLabel(welcome, image=logo_ctk, text="").pack(pady=(20, 10))
ctk.CTkLabel(welcome, text="🍰 Sweet Bakery", font=ctk.CTkFont(size=30, weight="bold"), text_color="#5A3E2B").pack()
ctk.CTkLabel(welcome, text="Enter your name", font=ctk.CTkFont(size=16)).pack(pady=10)
name_var = ctk.StringVar()
ctk.CTkEntry(welcome, textvariable=name_var, width=300, fg_color="#F7CC3E", text_color="#5A3E2B").pack(pady=5)

def go_to_menu():
    nm = name_var.get().strip()
    if not nm:
        messagebox.showwarning("Missing", "Please enter your name.")
        return
    order_state["customer"] = nm
    name_label.configure(text=f"Customer: {nm}")
    show(menu)

ctk.CTkButton(welcome, text="Next →", width=200, fg_color="#F7CC3E", text_color="#5A3E2B", command=go_to_menu).pack(pady=25)

# ---------------------- Page 2 ----------------------
menu = ctk.CTkFrame(root, fg_color="#FFF8E1", corner_radius=16)
if logo_ctk:
    ctk.CTkLabel(menu, image=logo_ctk, text="").pack(pady=(20, 10))
ctk.CTkLabel(menu, text="🍩 Bakery Menu", font=ctk.CTkFont(size=26, weight="bold"), text_color="#5A3E2B").pack()

grid = ctk.CTkFrame(menu, fg_color="transparent")
grid.pack(expand=True, pady=20)
r = c = 0
for n, f, p in ITEMS:
    fr = ctk.CTkFrame(grid, fg_color="#F8E8C0", corner_radius=10)
    fr.grid(row=r, column=c, padx=15, pady=15)
    try:
        if f.exists():
            img = Image.open(f).resize((100, 100))
            im = ctk.CTkImage(light_image=img, dark_image=img, size=(100, 100))
            ctk.CTkLabel(fr, image=im, text="").pack(pady=(10, 6))
    except:
        ctk.CTkLabel(fr, text="(no img)").pack(pady=(10, 6))
    ctk.CTkLabel(fr, text=f"{n}\n₹{p}", font=ctk.CTkFont(size=13, weight="bold")).pack(pady=(0, 10))
    c += 1
    if c == 4:
        c = 0
        r += 1
for i in range(4):
    grid.grid_columnconfigure(i, weight=1)

bottom = ctk.CTkFrame(menu, fg_color="transparent")
bottom.pack(pady=10)
ctk.CTkButton(bottom, text="← Back", width=140, fg_color="#B0BEC5", text_color="#263238", command=lambda: show(welcome)).pack(side="left", padx=10)
ctk.CTkButton(bottom, text="Continue →", width=200, fg_color="#F7CC3E", text_color="#5A3E2B", command=lambda: show(order)).pack(side="right", padx=10)

# ---------------------- Page 3 ----------------------
order = ctk.CTkFrame(root, fg_color="#FFF8E1", corner_radius=16)

header = ctk.CTkFrame(order, fg_color="#FFF8E1")
header.pack(fill="x", pady=(10, 5))
if logo_ctk:
    ctk.CTkLabel(header, image=logo_ctk, text="").pack(side="left", padx=10)
ctk.CTkLabel(header, text="🧁 Place Order", font=ctk.CTkFont(size=26, weight="bold"), text_color="#5A3E2B").pack(side="left", padx=10)
name_label = ctk.CTkLabel(header, text="Customer: Guest", font=ctk.CTkFont(size=14))
name_label.pack(side="right", padx=20)

split = ctk.CTkFrame(order, fg_color="transparent")
split.pack(expand=True, fill="both", padx=15, pady=10)
left = ctk.CTkFrame(split, fg_color="#FFF3D9", corner_radius=12)
right = ctk.CTkFrame(split, fg_color="#FFF3D9", corner_radius=12)
left.place(relx=0.0, rely=0, relwidth=0.68, relheight=1)
right.place(relx=0.70, rely=0, relwidth=0.29, relheight=1)

r = c = 0
for n, f, p in ITEMS:
    fr = ctk.CTkFrame(left, fg_color="#F8E8C0", corner_radius=10)
    fr.grid(row=r, column=c, padx=8, pady=8)
    ctk.CTkCheckBox(fr, text="", variable=order_state["selected"][n], fg_color="#F7CC3E").pack(anchor="ne", padx=5, pady=2)
    try:
        if f.exists():
            img = Image.open(f).resize((80, 80))
            im = ctk.CTkImage(light_image=img, dark_image=img, size=(80, 80))
            ctk.CTkLabel(fr, image=im, text="").pack()
    except:
        ctk.CTkLabel(fr, text="(no img)").pack()
    ctk.CTkLabel(fr, text=f"{n}\n₹{p}", font=ctk.CTkFont(size=12, weight="bold")).pack()
    qf = ctk.CTkFrame(fr, fg_color="transparent")
    qf.pack(pady=4)
    def dec(nn=n):
        if order_state["quantities"][nn] > 0:
            order_state["quantities"][nn] -= 1
            qty_labels[nn].configure(text=str(order_state["quantities"][nn]))
            update_summary()
    def inc(nn=n):
        order_state["quantities"][nn] += 1
        qty_labels[nn].configure(text=str(order_state["quantities"][nn]))
        update_summary()
    ctk.CTkButton(qf, text="-", width=28, command=dec, fg_color="#F7CC3E", text_color="#5A3E2B").grid(row=0, column=0, padx=2)
    lbl = ctk.CTkLabel(qf, text="0")
    lbl.grid(row=0, column=1, padx=3)
    qty_labels[n] = lbl
    ctk.CTkButton(qf, text="+", width=28, command=inc, fg_color="#F7CC3E", text_color="#5A3E2B").grid(row=0, column=2, padx=2)
    c += 1
    if c == 3:
        c = 0
        r += 1
for i in range(3):
    left.grid_columnconfigure(i, weight=1)

summary_label = ctk.CTkLabel(right, text="🧾 Invoice Summary", font=ctk.CTkFont(size=18, weight="bold"), text_color="#5A3E2B")
summary_label.pack(pady=(15, 5))
summary_text = ctk.CTkTextbox(right, height=400, width=300, fg_color="white")
summary_text.pack(padx=10, pady=10)

def update_summary():
    summary_text.delete("1.0", "end")
    summary_text.insert("end", "Item\tQty\tPrice\tTotal\n" + "-"*45 + "\n")
    total = 0
    for n, _, p in ITEMS:
        if order_state["selected"][n].get() and order_state["quantities"][n] > 0:
            t = p * order_state["quantities"][n]
            total += t
            summary_text.insert("end", f"{n}\t{order_state['quantities'][n]}\t₹{p}\t₹{t}\n")
    summary_text.insert("end", "-"*45 + f"\nGrand Total:\t\t₹{total}\n")
    return total
update_summary()

# ---------------------- Buttons ----------------------
buttons = ctk.CTkFrame(order, fg_color="transparent")
buttons.pack(fill="x", pady=(5, 15))

def confirm_order():
    items = []
    grand = update_summary()
    for n, _, p in ITEMS:
        q = order_state["quantities"][n]
        if order_state["selected"][n].get() and q > 0:
            cost = p * q
            items.append((n, q, p, cost, grand))
    if not items:
        messagebox.showwarning("Empty", "Please select some items.")
        return
    customer = order_state["customer"]
    append_owner_record(customer, grand)
    saved = append_customer_file(customer, items)
    messagebox.showinfo("Saved", f"Invoice saved for {customer} at:\n{saved}\nOwner record updated!")

ctk.CTkButton(buttons, text="← Back", width=200, fg_color="#B0BEC5", text_color="#263238", command=lambda: show(menu)).pack(side="left", padx=40)
ctk.CTkButton(buttons, text="Confirm Order", width=240, fg_color="#81C784", text_color="#2E7D32", command=confirm_order).pack(side="right", padx=40)

# ---------------------- Start ----------------------
show(welcome)
root.mainloop()
